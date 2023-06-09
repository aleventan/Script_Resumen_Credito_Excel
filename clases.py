import re

from leerPDF import abrirPDF
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter


def siguiente_mes(mes_actual):
    meses = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
    indice_mes_actual = meses.index(mes_actual.lower())
    indice_siguiente_mes = (indice_mes_actual + 1) % 12
    return meses[indice_siguiente_mes]

class tablaExcel:
    def __init__(self):
        self.hoja = None
        self.tabla = None
        self.wb = None

    def tomarDatosPDF(self, nombre_archivo_xls, tipo_dato=0):
        # Importa los datos del pdf
        self.datos = abrirPDF(nombre_archivo_xls)
        if tipo_dato == 0 or tipo_dato == 1: # Consumos
            self.consumos = self.datos[0]
        elif tipo_dato == 2: # Impuestos
            self.impuestos = self.datos[1]
        elif tipo_dato == 3: # Pagos
            self.pagos = self.datos[2]
        self.fecha_vto_completa = self.datos[3].lower()
        self.fecha_vto = self.datos[3][3:].lower()

    def abrirExcel(self, nombre_archivo, hoja):
        # Toma el nombre del archivo y la hoja
        self.nombre_archivo = nombre_archivo
        self.wb = openpyxl.load_workbook(nombre_archivo)
        self.hoja = self.wb[hoja]


    def maxRowCol(self):
        # Creo la función para que cada vez que se necesite se recuente la cantidad de columnas y filas
        self.max_column, self.max_row = self.hoja.max_column, self.hoja.max_row
        return self.max_column, self.max_row

    def formatoInicial(self, tabla):
        # Borra la tabla como objeto, no los datos, ya que no se abre correctamente. Luego se arma de nuevo la tabla como objeto
        self.maxRowCol()
        if self.hoja.tables:
            del self.hoja.tables[tabla]
        # Muestra todas las filas
        for fila in range(self.max_row):
            self.hoja.row_dimensions[fila].hidden = False

        # Sacar los bordes
        borde_por_defecto = Border()
        for fila in range(1, self.max_row + 1):
            for columna in range(1, self.max_column + 1):
                letra_columna = get_column_letter(columna)
                celda = self.hoja['{}{}'.format(letra_columna, fila)]
                celda.border = borde_por_defecto



    def recorrerDatosExistentes(self, desde_fila=2, cant_col_iterar=2):
        try:
            self.maxRowCol()
            # Creo un objeto de lista para guardar los datos ya existentes
            self.tabla_guardada = []
            # Iterar sobre cada fila desde la 2da y guardalas en una lista para comparar despues
            for fila in range(desde_fila, self.max_row + 1):
                # Iterar sobre cada celda de la fila actual y guardar en una lista los valores de la fila
                fila_guardar = []
                for columna in range(cant_col_iterar):
                    # Obtener el valor de la celda actual
                    valor_celda = self.hoja.cell(row=fila + 1, column=columna + 1).value
                    if valor_celda is None:
                        break
                    else:
                        fila_guardar.append(valor_celda)
                if fila_guardar:
                    if self.hoja.title == 'Cuotas':
                        fila_guardar.append(fila + 1)
                    else:
                        fila_guardar.insert(0, fila + 1)
                    self.tabla_guardada.append(fila_guardar)
            return self.tabla_guardada
        except Exception as e:
            # Obtener información detallada de la excepción y la pila de llamadas
            traceback_info = traceback.format_exc()

            # Mostrar el nombre de la función y el número de línea donde se produjo el error
            print("Mensaje de error:", str(e))

            # Obtener la pila de llamadas
            stack_trace = traceback.extract_tb(e.__traceback__)
            # Obtener el último elemento de la pila de llamadas
            last_call = stack_trace[-1]

            # Mostrar el nombre del archivo, nombre de la función y número de línea
            print("Se produjo un error en el archivo:", last_call.filename)
            print("En la función:", last_call.name)
            print("Número de línea:", last_call.lineno)

    def mesActual(self, col_inicio):
        # la variable mes_actual indica la columna donde se empieza a cargar cada importe. Arranca en la columna 4 si no hay datos.
        # Si los hay, desde la que corresponde
        self.maxRowCol()
        if self.hoja.title == 'Cuotas':
            self.mes_actual = None
            for c in range(self.max_column):
                valor_celda = self.hoja.cell(row=1, column=c + 1).value
                if valor_celda == self.fecha_vto:
                    self.mes_actual = c + 1
                    break

            if self.mes_actual is None:
                self.mes_actual = col_inicio
        else:
            self.mes_actual = None
            encontrado = 'N'
            for c in range(self.max_column + 1):
                valor_celda = self.hoja.cell(row=1, column=c + 1).value
                if valor_celda == self.fecha_vto:
                    self.mes_actual = c + 1
                    encontrado = 'S'
                    break
                if valor_celda == 'Columna1':
                    self.mes_actual = col_inicio
                    self.hoja.cell(row=1, column=c + 1).value = self.fecha_vto
                    encontrado = 'S'
                    break

            if encontrado == 'N':
                self.mes_actual = self.max_column + 1
                self.hoja.cell(row=1, column=self.max_column + 1).value = self.fecha_vto
            #     if self.max_column == col_inicio:
            #         print(self.mes_actual)
            #         if self.mes_actual is None:
            #             self.mes_actual = col_inicio
            #     else:
                # self.hoja.cell(row=1, column=self.max_column + 1).value = self.fecha_vto

        return self.mes_actual
    
    def agregarConsumosCuotas(self, row=3):
        self.recorrerDatosExistentes()
        self.mesActual(4)
        self.maxRowCol()
        row = row  # Empezar en la TERCER fila para evitar sobrescribir los encabezados
        # Contador para ir agregando los datos en la fila necesaria
        suma_contador = 0
        self.hoja.cell(row=1, column=3).value = 'Cant.Cuotas'
        # Recorre la lista de consumos
        for consumo in self.consumos:
            # Si el consumo no es cuota, sigue
            if consumo['cuota'] != '':
                # Si hay una tabla creada, revisa si encuentra el item y si lo hace, lo setea en S, indicandole en que fila se encontraba
                if self.tabla_guardada:
                    encontro_item = 'N'
                    for item in self.tabla_guardada:
                        if item[0] == consumo['id_consumo'] and item[1][:7] == consumo['consumo'][:7]:
                            encontro_item = 'S'
                            nro_fila_cuota = item[2] + suma_contador
                            break
                # Si no hay tabla, procede a setearlo en N para cargarlo despues
                else:
                    encontro_item = 'N'

                if encontro_item == 'S':
                    # Actualiza el número de cuota actual
                    self.hoja.cell(row=nro_fila_cuota, column=3).value = consumo['cuota']

                else:
                    # agrego los datos nuevos
                    self.hoja.insert_rows(3)
                    suma_contador += 1
                    self.hoja.cell(row=row, column=1).value = consumo['id_consumo']
                    self.hoja.cell(row=row, column=2).value = consumo['consumo']
                    self.hoja.cell(row=row, column=3).value = consumo['cuota']

                    # obtener el número de cuotas restantes
                    cuotas_restantes = int(consumo['cuota'][3:]) - int(consumo['cuota'][:2]) + 1
                    # escribir el importe en cada celda correspondiente
                    for i in range(cuotas_restantes):
                        self.hoja.cell(row=row, column=i + self.mes_actual, value=consumo['importe'])

        # Eliminar la última fila
        self.hoja.delete_rows(self.hoja.max_row)


    def agregarUnPago(self):
        self.mesActual(col_inicio=2)
        self.maxRowCol()
        self.recorrerDatosExistentes(desde_fila=1)
        row = 3  # Empezar en la TERCER fila para evitar sobrescribir los encabezados

        # Filtrar los elementos con 'cuota' vacía y armar una lista de ellos
        filtrados = [con for con in self.consumos if con['cuota'] == '']

        # Crea un diccionario de los pagos repetidos y los suma
        agrupados = {}
        for c in filtrados:
            consumo = c['consumo']
            importe = c['importe']
            if consumo in agrupados:
                agrupados[consumo] += importe
            else:
                agrupados[consumo] = importe

        # Agrega el consumo
        for consumo, importe_total in agrupados.items():
            # Si la tabla y el consumo existen, marca al item, sino lo deja como que no lo encontró
            if self.tabla_guardada:
                encontro_item = 'N'
                for item in self.tabla_guardada:
                    if item[1] == consumo:
                        encontro_item = 'S'
                        break
            else:
                encontro_item = 'N'

            # Si encuentra al item en la tabla guardada, comprueba en donde se encuentra en toda la tabla y agrega el importe en el mes actual.
            if encontro_item == 'S':
                for fila in range(1, self.max_row + 1):
                    tipo_consumo = self.hoja.cell(row=fila, column=1).value  # Supongamos que el consumo se encuentra en la columna 1
                    if tipo_consumo == consumo:
                        self.hoja.cell(row=fila, column=self.mes_actual).value = importe_total
                        break
            # Si no lo encuentra, inserta una fila (en la 3) y agrega el consumo y el total del importe sumado anteriormente
            else:
                self.hoja.insert_rows(row)
                self.hoja.cell(row=row, column=1).value = consumo
                self.hoja.cell(row=row, column=self.mes_actual).value = importe_total

        # Eliminar la fila si no tiene consumo
        for f in range(1, self.hoja.max_row):
            valor_celda = self.hoja.cell(row=f, column=1).value
            if valor_celda is None:
                self.hoja.delete_rows(self.hoja.max_row)



    def agregarImpuestos(self):
        self.mesActual(col_inicio=2)
        self.maxRowCol()
        self.recorrerDatosExistentes(desde_fila=1)
        row = 3  # Empezar en la TERCER fila para evitar sobrescribir los encabezados

        # Agrega el impuesto a la tabla
        for impuesto in self.impuestos:
            # Si la talba y el nombre del impuesto existen, lo guarda para luego anotar el importe, sino marca que no existe
            if self.tabla_guardada:
                encontro_item = 'N'
                for item in self.tabla_guardada:
                    if item[1] == impuesto['consumo']:
                        encontro_item = 'S'
                        break
            else:
                encontro_item = 'N'

            # Si existe el impuesto, agrega el importe en el mes correspondiente
            if encontro_item == 'S':
                for fila in range(1, self.max_row + 1):
                    tipo_impuesto = self.hoja.cell(row=fila, column=1).value  # Supongamos que el consumo se encuentra en la columna 1
                    if tipo_impuesto == impuesto['consumo']:
                        self.hoja.cell(row=fila, column=self.mes_actual).value = impuesto['importe']
                        break

            # Sino, agrega la fila 3 y ahí lo anota
            else:
                self.hoja.insert_rows(row)
                self.hoja.cell(row=row, column=1).value = impuesto['consumo']
                self.hoja.cell(row=row, column=self.mes_actual).value = impuesto['importe']

        # # Eliminar la fila si no tiene consumo
        for f in range(1, self.hoja.max_row):
            valor_celda = self.hoja.cell(row=f, column=1).value
            if valor_celda is None:
                self.hoja.delete_rows(self.hoja.max_row)


    def agregarPagos(self):
        self.mesActual(col_inicio=2)
        self.maxRowCol()
        self.recorrerDatosExistentes(desde_fila=1)
        row = 3  # Empezar en la TERCER fila para evitar sobrescribir los encabezados

        # Filtrar los elementos con 'SU PAGO EN PESOS', por si aparece otro tipo de pago
        filtrados = [con for con in self.pagos if con['consumo'] == 'SU PAGO EN PESOS']

        # Crea un diccionario con la suma de todos los pagos
        agrupados = {}
        for c in filtrados:
            consumo = c['consumo']
            importe = c['importe']
            if consumo in agrupados:
                agrupados[consumo] += importe
            else:
                agrupados[consumo] = importe

        # Agrega los pagos a la tabla, si es el primero crea el item
        for consumo, importe_total in agrupados.items():
            if self.tabla_guardada:
                encontro_item = 'N'
                for item in self.tabla_guardada:
                    if item[1] == consumo:
                        encontro_item = 'S'
                        break
            else:
                encontro_item = 'N'

            if encontro_item == 'S':
                for fila in range(1, self.max_row + 1):
                    tipo_pago = self.hoja.cell(row=fila, column=1).value  # Supongamos que el consumo se encuentra en la columna 1
                    if tipo_pago == consumo:
                        self.hoja.cell(row=fila, column=self.mes_actual).value = importe_total
                        break

            else:
                self.hoja.insert_rows(row)
                self.hoja.cell(row=row, column=1).value = consumo
                self.hoja.cell(row=row, column=self.mes_actual).value = importe_total

        # # Eliminar la fila si no tiene consumo
        for f in range(1, self.hoja.max_row):
            valor_celda = self.hoja.cell(row=f, column=1).value
            if valor_celda is None:
                self.hoja.delete_rows(self.hoja.max_row)



    def agregarMesActual(self):
        self.maxRowCol()

        row = 2

        # Sacar los bordes
        borde_por_defecto = Border()
        for fila in range(1, self.max_row + 1):
            for columna in range(1, self.max_column + 1):
                letra_columna = get_column_letter(columna)
                celda = self.hoja['{}{}'.format(letra_columna, fila)]
                celda.border = borde_por_defecto

        # Eliminar datos del mes anterior
        formato_rango_consumos = 'A2:D' + str(self.max_row)
        rango_consumos = self.hoja[formato_rango_consumos]

        formato_rango_impuestos = 'H2:I' + str(self.max_row)
        rango_impuestos = self.hoja[formato_rango_impuestos]

        # Iterar sobre cada celda del rango y asignar un valor vacío
        for fila in rango_consumos:
            for celda in fila:
                celda.value = None

        for fila in rango_impuestos:
            for celda in fila:
                celda.value = None


        # Elimino formato de tabla
        # Crear una copia del diccionario self.hoja.tables
        copiatablas = dict(self.hoja.tables)

        # Iterar sobre la copia y eliminar los elementos originales
        for tabla in copiatablas:
            del self.hoja.tables[tabla]

        # Creo una fila en base a la primera para que se modifique solo en esta tabla
        row_consumo = row
        # Agrego TODOS los consumos del mes
        for consumo in self.consumos:
            for fila in range(1, len(self.consumos)):
                self.hoja.cell(row=row_consumo, column=1).value = consumo['fecha']
                self.hoja.cell(row=row_consumo, column=2).value = consumo['consumo']
                self.hoja.cell(row=row_consumo, column=3).value = consumo['cuota']
                self.hoja.cell(row=row_consumo, column=4).value = consumo['importe']
                row_consumo += 1
                break

        # Creo una fila en base a la primera para que se modifique solo en esta tabla
        row_impuesto = row
        # Agrego todos los impuestos
        for impuesto in self.impuestos:
            for fila in range(0, len(self.impuestos)):
                self.hoja.cell(row=row_impuesto, column=8).value = impuesto['consumo']
                self.hoja.cell(row=row_impuesto, column=9).value = impuesto['importe']
                row_impuesto += 1
                break

        pagos_agrupados = 0
        # Filtrar los elementos con pagos
        filtrados = [con for con in self.pagos if con['consumo'] == 'SU PAGO EN PESOS']

        # Sumo el total de los pagos realizados
        for c in filtrados:
            importe = c['importe']
            pagos_agrupados += importe


        self.hoja.cell(row=row, column=13).value = pagos_agrupados

        ## Estilo tabla ##
        # Creo los rangos de las tablas y luego los agrego a un diccionario para que aplique el formato a las tablas #
        rango_consumos = 'A1:D' + str(row_consumo - 1)
        rango_impuestos = 'H1:I' + str(row_impuesto - 1)
        rango_pagos = 'M1:M2'
        tablas_mes = {'Table_Consumos_mes': rango_consumos, 'Table_Impuestos_mes': rango_impuestos, 'Table_Pagos_mes': rango_pagos}
        # Definir el estilo de borde
        borde = Border(left=Side(border_style='thin'),
                       right=Side(border_style='thin'),
                       top=Side(border_style='thin'),
                       bottom=Side(border_style='thin'))

        # Recorro el diccionario de tablas
        for tabla, rango in tablas_mes.items():
            # Creo la tabla como objeto
            tab = Table(displayName=tabla, ref=rango)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name='TableStyleLight11', showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            self.hoja.add_table(tab)
            ref = self.hoja.tables[tabla].ref
            rango_tabla = self.hoja[ref]  # Obtener el rango de celdas de la tabla

            # Agregar borde a toda la tabla
            for fila in rango_tabla:
                for celda in fila:
                    celda.border = borde


        # Agrego la Fecha de vencimiento y le doy formato a eso
        self.hoja.cell(row=9, column=13).value = self.fecha_vto_completa
        celda_m8 = self.hoja['M8']
        celda_m9 = self.hoja['M9']
        celda_m8.border = borde
        celda_m9.border = borde

        # Crear un objeto NumberFormat con el formato de pesos argentinos y los aplica a todas las columnas necesarias
        pesos_formato = '$ #,##0.00'
        columnas = ['D', 'I', 'M']
        for columna in columnas:
            for fila in range(1, self.hoja.max_row + 1):
                celda = self.hoja[columna + str(fila)]
                celda.number_format = pesos_formato

    def agregarTitulos(self, col_inicio=3):
        # Agrego titulo a donde sea necesario

        self.maxRowCol()
        # Si la hoja es Cuotas
        if self.hoja.title =='Cuotas':
            for z in range(col_inicio, self.max_column - 1):
                valor = self.hoja.cell(row=1, column=z + 1).value
                # Si el valor de la primer columna donde van los meses es vacio
                if z == col_inicio and valor == 'Columna1':
                    self.hoja.cell(row=1, column=z + 1).value = self.fecha_vto
                    valor = self.hoja.cell(row=1, column=z + 1).value
                # Si el valor tiene algo
                if valor is not None:
                    # Separa los meses y recorre para ver cual es el siguiente, sumando uno para el año cuando empieza ene
                    mes_pasado = valor[:3]
                    anio_corriente = int(valor[4:])
                    prox_mes = siguiente_mes(mes_pasado)
                    if prox_mes == 'ene':
                        anio_corriente = int(anio_corriente) + 1
                    self.hoja.cell(row=1, column=z + 2).value = prox_mes + '-' + str(anio_corriente)

        # Como el resto de las hojas el titulo no se necesita saber el siguiente mes, se ejecuta esto
        else:
            for z in range(col_inicio, self.max_column + 1):
                valor = self.hoja.cell(row=1, column=z).value
                # Si el valor esta vacio, agrega el mes
                if z == col_inicio and valor == 'Columna1':
                    self.hoja.cell(row=1, column=z).value = self.fecha_vto
                    valor = self.hoja.cell(row=1, column=z).value

                #Si cumple con el patron, no hace mada, sino agrega la fecha_vto
                patron = r'^[a-z]{3}-\d{2}$'
                if isinstance(valor, str) and not re.match(patron, valor, re.IGNORECASE):
                    self.hoja.cell(row=1, column=z).value = self.fecha_vto





    def sumaTotal(self, col_inicio=4, fila_inicio=3):
        self.maxRowCol()

        # Le da formato a la celda
        fuente = Font(bold=True, size=12)

        # Escribir la suma de cada columna en la fila 2 de la hoja
        for columna in range(col_inicio, self.max_column + 1):
            letra_columna = get_column_letter(columna)
            # Genero la fórmula para que quede en la celda, y si se modifica algun valor 'a mano', modifica la suma
            formula = '=SUM('+ letra_columna + '3:' + letra_columna + str(self.max_row) + ')'
            if self.hoja != 'Cuota':
                if self.hoja.cell(row=2, column=1).value is not None:
                    self.hoja.insert_rows(2)
            self.hoja.cell(row=2, column=columna, value=formula)
            # Obtiene la celda para aplicar el formato
            celda = self.hoja[get_column_letter(columna) + str(2)]
            celda.font = fuente



    def rangosTabla(self):
        # Definir el rango de celdas
        self.maxRowCol()
        start_cell = self.hoja.cell(row=1, column=1)
        last_cell = self.hoja.cell(row=self.max_row, column=self.max_column)
        self.range_str = f"{start_cell.coordinate}:{last_cell.coordinate}"
        return self.range_str
    
    def formatoFinal(self, tabla, estiloTabla="TableStyleLight11", col_inicio=4, fila_inicio=3):
        self.maxRowCol()
        self.rangosTabla()

        # Si la hoja es Cuotas
        if self.hoja.title == 'Cuotas':
            self.mesActual(col_inicio=4)
            # Genero el nombre y los rangos de la tabla
            tab = Table(displayName=tabla, ref=self.range_str)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name=estiloTabla, showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            self.hoja.add_table(tab)

            self.hoja.row_dimensions[1].auto_size = True

            # Agregar borde a toda la tabla
            # Definir el estilo de borde
            borde = Border(left=Side(border_style='thin'),
                           right=Side(border_style='thin'),
                           top=Side(border_style='thin'),
                           bottom=Side(border_style='thin'))

            # Aplicar el estilo de borde a cada celda en el rango
            for fila in range(1, self.max_row + 1):
                for columna in range(1, self.max_column + 1):
                    letra_columna = get_column_letter(columna)
                    celda = self.hoja['{}{}'.format(letra_columna, fila)]
                    celda.border = borde

                    self.hoja.column_dimensions[letra_columna].auto_size = True
                self.hoja.row_dimensions[fila].auto_size = True

            # Marco borde mes actual
            borde_actual = Border(left=Side(border_style='medium'),
                                  right=Side(border_style='medium'),
                                  top=Side(border_style='medium'),
                                  bottom=Side(border_style='medium'))

            # Establecer el color de relleno
            relleno_mes_previo = PatternFill(start_color='B3C6E7', end_color='B3C6E7', fill_type='solid')
            relleno_mes_actual = PatternFill(start_color='FF2c6fe4', end_color='FF2c6fe4', fill_type='solid')

            # Iterar a través de todas las filas en la columna y establecer el color de fondo para cada celda
            for col in range(col_inicio, self.mes_actual + 1):
                for fila in range(fila_inicio, self.max_row + 1):
                    celda = self.hoja['{}{}'.format(get_column_letter(col), fila)]

                    if col == self.mes_actual:
                        celda.fill = relleno_mes_actual
                        celda.border = borde_actual
                        # print(col, self.mes_actual)
                    else:
                        celda.fill = relleno_mes_previo

            # Crear un objeto NumberFormat con el formato de pesos argentinos
            pesos_formato = '$ #,##0.00'

            # Aplicar el formato a las celdas deseadas
            for fila in self.hoja.iter_rows(min_row=2, max_row=self.max_row, min_col=4, max_col=self.max_column):
                for celda in fila:
                    celda.number_format = pesos_formato
        else:
            self.mesActual(col_inicio=2)
            # Toma la tabla para editar los rangos
            tab = Table(displayName=tabla, ref=self.range_str)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name=estiloTabla, showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            self.hoja.add_table(tab)

            # Agregar borde a toda la tabla
            # Definir el estilo de borde
            borde = Border(left=Side(border_style='thin'),
                           right=Side(border_style='thin'),
                           top=Side(border_style='thin'),
                           bottom=Side(border_style='thin'))

            # Aplicar el estilo de borde a cada celda en el rango
            for fila in range(1, self.max_row + 1):
                for columna in range(1, self.max_column + 1):
                    letra_columna = get_column_letter(columna)
                    celda = self.hoja['{}{}'.format(letra_columna, fila)]
                    celda.border = borde
                    self.hoja.column_dimensions[letra_columna].auto_size = True
                self.hoja.row_dimensions[fila].auto_size = True

            #Marco borde mes actual
            borde_actual = Border(left=Side(border_style='medium'),
                           right=Side(border_style='medium'),
                           top=Side(border_style='medium'),
                           bottom=Side(border_style='medium'))

            # Establecer el color de relleno
            relleno_mes_previo = PatternFill(start_color='B3C6E7', end_color='B3C6E7', fill_type='solid')
            relleno_mes_actual = PatternFill(start_color='FF2c6fe4', end_color='FF2c6fe4', fill_type='solid')

            # Iterar a través de todas las filas en la columna y establecer el color de fondo para cada celda
            for col in range(col_inicio, self.mes_actual + 1):
                for fila in range(fila_inicio, self.max_row + 1):
                    celda = self.hoja['{}{}'.format(get_column_letter(col), fila)]
                    if col == self.mes_actual:
                        celda.fill = relleno_mes_actual
                        celda.border = borde_actual
                        # print(col, self.mes_actual)
                    else:
                        celda.fill = relleno_mes_previo

            # Crear un objeto NumberFormat con el formato de pesos argentinos
            pesos_formato = '$ #,##0.00'

            # Aplicar el formato a las celdas deseadas
            for fila in self.hoja.iter_rows(min_row=2, max_row=self.max_row, min_col=2, max_col=self.max_column):
                for celda in fila:
                    celda.number_format = pesos_formato



                
    def ocultarColFilas(self, fila_inicio=3):
        self.maxRowCol()
        # ocultar las columnas. Se deja visible una anterior a la actual
        if self.hoja.title == 'Cuotas':
            if self.mes_actual >= 6:
                self.hoja.column_dimensions.group(start=get_column_letter(4),
                                                  end=get_column_letter(self.mes_actual - 2), hidden=True)
        else:
            if self.mes_actual > 3:
                self.hoja.column_dimensions.group(start='B', end=get_column_letter(self.mes_actual - 2), hidden=True)

        # Para agrupar filas, arma una lista con las filas vacias
        filas_vacias = []

        for fila in range(fila_inicio, self.max_row + 1):
            celda_primera = self.hoja.cell(row=fila, column=self.mes_actual - 1).value
            celda_segunda = self.hoja.cell(row=fila, column=self.mes_actual).value

            if self.hoja.title == 'Cuotas':
                if celda_primera is None and celda_segunda is None:
                    filas_vacias.append(fila)
            else:
                if celda_segunda is None:
                    filas_vacias.append(fila)


        # Ordena la lista y arma una lista y dentro otras listas de filas contiguas
        grupos_contiguos = []
        grupo_actual = []
        # print(filas_vacias)
        filas_vacias.sort()
        for fila in filas_vacias:
            if not grupo_actual or fila == grupo_actual[-1] + 1:
                grupo_actual.append(fila)
            else:
                grupos_contiguos.append(grupo_actual)
                grupo_actual = [fila]

        if grupo_actual:
            grupos_contiguos.append(grupo_actual)

        # Agrupa segun una o mas
        for grupo in grupos_contiguos:
            if len(grupo) > 1:
                self.hoja.row_dimensions.group(start=grupo[0], end=grupo[-1], hidden=True)
            else:
                self.hoja.row_dimensions.group(start=grupo[0], end=grupo[0], hidden=True)


    def guardarTabla(self):
        self.wb.save(self.nombre_archivo)

    def cerrarExcel(self):
        # Cerrar el libro de trabajo
        self.wb.close()


if __name__ == '__main__':
    nueva_tabla = tablaExcel()
    # nueva_tabla.tomarDatosPDF()
    # nueva_tabla.abrirExcel("Detalle Visa Nacion.xlsx", 'Cuotas')
    # # valor = nueva_tabla.obtener_valor_celda(5, 1)
    # recorerDatos = nueva_tabla.recorrerDatosExistentes()
    # mes_actual = nueva_tabla.mesActual()
    # for r in recorerDatos:
    #     print(r)
    # nueva_tabla.agregarConsumosCuotas()
    # nueva_tabla.guardarTabla()
    # nueva_tabla.cerrarExcel()